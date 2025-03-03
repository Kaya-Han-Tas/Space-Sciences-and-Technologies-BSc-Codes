from scipy.optimize import curve_fit
from matplotlib import pyplot as plt
import math
import numpy as np
import openpyxl
import sys

#Excel Dosyasından data çekilecektir
#Açtığımız andaki dosya "Workbook" olarak geçer ve üzerine yazılabilir.
#Excel dosyamıza yazabilmek için Dosyanın kapatılması gerekmekte olup, açık olmamalıdır! 
#Aksi takdirde hata alınacaktır.
datafile_open=openpyxl.load_workbook(r"D:\UBT 4. SINIF\UBT\Uzay Fiziğinde Seçilmiş Konular\Ödevler\V0410 Aur.xlsx")

#.active ile açılan dosya bir "active sheet" haline getirilir ve buradan datalar çekilebilir.
datafile=datafile_open.active

#Data Dosyamızdan kullanılacak T0 ve P değeri çekilir.
T0=float((datafile.cell(row=2,column=4).value)) #Başlangıç Zamanı
P=float((datafile.cell(row=2,column=5).value)) #Periyot

#Buradan itibaren webden alınan değerler "obs" ve hesaplanan değerler "calc" ile gösterilmektedir.
#Bununla beraber diğer Dataları da çekerek kendi listelerine aktarıyoruz. (Webden Alınan Veriler)
Evalues_obs=[]
Tvalues_obs=[]
OCvalues_obs=[]
x=(datafile.max_row)-1
y=1

while y<=x:
    y+=1
    Edata=(datafile.cell(row=y,column=1).value)
    Evalues_obs.append(Edata)
    OCdata=(datafile.cell(row=y,column=2).value)
    OCvalues_obs.append(OCdata)
    Tdata=(datafile.cell(row=y,column=3).value)
    Tvalues_obs.append(Tdata)

#T değerleri ve OC değerleri Excel Dosyasında "Metin" olarak yazıldığından bunları sayıya çevirmemiz gerekir.
Tvalues_obs=[float (i) for i in Tvalues_obs]
OCvalues_obs=[float(i) for i in OCvalues_obs]

#Şimdi datalarımızı array'e çevirerek işimizi kolaylaştırabiliriz.
Tvalues_obs=np.asarray(Tvalues_obs)
Evalues_obs=np.asarray(Evalues_obs)
OCvalues_obs=np.asarray(OCvalues_obs)

#İlk olarak kendimiz E değerlerini hesaplıyoruz.
#Bunun için E=(T-T0)/P formülü kullanılacaktır.
#Buna dikkat edilerek işlem gerçekleştirilir.
#Ayrıca E değerlerinin yuvarlanmasında bir kural üzerinden ilerlememiz gerekir.
#Bunun için bulduğumuz Hesaplanan "Yuvarlanmış" E değeri ile Hesaplanan E değeri arasındaki farkın mutlak değerinin 0.3'ten büyük olduğunu varsayıyoruz.

#Hesaplanan "Yuvarlanmış" Değer, Hesaplanan Değerden küçükse yani ikisinin farkı negatifse Yuvarlanan değerden 0.5 ÇIKARILIR.
#Hesaplanan "Yuvarlanmış" Değer, Hesaplanan Değerden büyükse yani ikisinin farkı pozitifse Yuvarlanan değere 0.5 EKLENİR.

#Hesaplanan "Yuvarlanmış" E değeri ile Hesaplanan E değeri arasındaki farkın mutlak değerinin 0.3'ten küçük olması durumunda yuvarlanan değer doğru kabul edilir.
#Buradan yola çıkarak E değerlerini doğru yuvarlanmış biçimleriyle elde ederiz.
Evalues_calc=[]
Evalues_calc_original=[]
for i in range(len(Evalues_obs)):
    E_calc=(Tvalues_obs[i]-T0)/P
    Evalues_calc_original.append(E_calc)
    
    if abs(E_calc-round(E_calc))>0.3:
        if (E_calc-round(E_calc))<0:
            Evalues_calc.append(round(E_calc)-0.5)
        
        else:
            Evalues_calc.append(round(E_calc)+0.5)
    
    else:
        Evalues_calc.append(round(E_calc))

#Şimdi bulunan E değerlerini Excel Dosyamıza yazmamız faydalı olacaktır.
#Bunun için de aşağıdaki kod kullanılır.
#Burada artı olarak bulunan değerleri "Metin" olarak Excel dosyasına yazmak faydalı olacaktır.

#i=0 olacağından ve 1. satırda başlık bulunduğundan sayılar excel dosyasına 2. satırdan itibaren yazılmalıdır.
#Bu nedenle i+2. satırdan itibaren değerleri Excel dosyasına yazmaktayız.

#Siteden Alınan E değerleri ile Hesaplanan E değerleri arası farkı da Excel Dosyasına yazdırarak kontrol yapıyoruz.
#Burada "Secondary" tutulmalardan elde edilen E değerleri arası fark çıkacaktır, buradan Secondary tutulma olduğu anlaşılabilir!

#Hesaplanan E Değerlerinin yuvarlanmamış durumlarını da kontrol için Excel Dosyasına yazdırıyoruz.

for i in range(len(Evalues_calc)):
    datafile.cell(row=1,column=7).value="Siteden Alınan E değerleri" #Başlık
    datafile.cell(row=i+2, column=7).value=(Evalues_obs[i]) #Değerler
    
    datafile.cell(row=1,column=8).value="Hesaplanan E Değerleri" #Başlık
    datafile.cell(row=i+2, column=8).value=(Evalues_calc[i]) #Değerler
    
    datafile.cell(row=1,column=9).value="Site ve Hesaplanan E değeri arası fark" #Başlık
    datafile.cell(row=i+2, column=9).value=(Evalues_obs[i]-Evalues_calc[i]) #Değerler
    
    datafile.cell(row=1,column=10).value="Hesaplanan E Değerleri (Yuvarlanmamış)" #Başlık
    datafile.cell(row=i+2, column=10).value=(Evalues_calc_original[i]) #Değerler

#İkinci olarak O-C Hesabımızı yapıyoruz.
#Bunun için T=T0+EP formülü kullanılacaktır.
#Bu formülde bulunan E değerleri Hesapladığımız E değerleridir!

#Bu formülden elde edilen değerler "Calculated" Zamanlardır.
#Siteden aldıklarımız ise "Observed" Zamanlardır.
#Bu iki zaman farkı da bize O-C değerlerini vermelidir.

#İlk olarak Calculated yani Hesaplanan Zaman değerlerini T=T0+EP formülü ile elde ediyoruz.
#Burada T0 değeri primary veya secondary tutulma olmasına göre belirlenmelidir!
#Bunu dikkate alarak Bulduğumuz T değerlerini ayrıca Excel Dosyasına yazdırıyoruz.
Tvalues_calc=[]
for i in range(len(Tvalues_obs)):
    T_calc=T0+(Evalues_calc[i]*P)
    Tvalues_calc.append(T_calc)
    
    datafile.cell(row=1,column=13).value="Hesaplanan T Değerleri (Calculated Time)" #Başlık
    datafile.cell(row=i+2, column=13).value=(T_calc) #Değerler
    
    datafile.cell(row=1, column=12).value="Siteden Alınan T Değerleri (Observed Time)"
    datafile.cell(row=i+2, column=12).value=(Tvalues_obs[i])

#Hesapladığımız Zaman değerleri ve Websitesinden aldığımız zaman değerlerini kullanarak O-C hesaplanır.
#Excel Dosyasına bu değer de yazdırılır.
OCvalues_calc=[]
for i in range(len(Tvalues_obs)):
    OC_calc=Tvalues_obs[i]-Tvalues_calc[i]
    OCvalues_calc.append(OC_calc)
    datafile.cell(row=1,column=14).value="Hesaplanan O-C Değerleri" #Başlık
    datafile.cell(row=i+2, column=14).value=(OC_calc) #Değerler
    
#Şimdi bu O-C değerlerinin Sitedeki O-C değerleri ile karşılaştırmasını Excel Dosyamıza yazmamız faydalı olacaktır.
#Bunun için de aşağıdaki kod kullanılır.
#Burada artı olarak bulunan değerleri "Metin" olarak Excel dosyasına yazmak faydalı olacaktır.

#Yine E için yaptığımız gibi Siteden alınan O-C değerleri ile Hesapladığımız O-C değerleri arası farkı Excel Dosyasına yazdırıyoruz.

#i=0 olacağından ve 1. satırda başlık bulunduğundan sayılar excel dosyasına 2. satırdan itibaren yazılmalıdır.
#Bu nedenle i+2. satırdan itibaren değerleri Excel dosyasına yazmaktayız.
for i in range(len(OCvalues_calc)):
    datafile.cell(row=1,column=16).value="Sitedeki O-C Değerleri" #Başlık
    datafile.cell(row=i+2, column=16).value=(OCvalues_obs[i]) #Değerler
    
    datafile.cell(row=1,column=17).value="Hesaplanan O-C Değerleri" #Başlık
    datafile.cell(row=i+2, column=17).value=str(round(OCvalues_calc[i],5)) #Değerler
    
    datafile.cell(row=1,column=18).value="Site ve Hesaplanan O-C değeri arası fark" #Başlık
    datafile.cell(row=i+2, column=18).value=(round(OCvalues_calc[i],5)-OCvalues_obs[i]) #Değerler

#Şimdi bulduğumuz O-C değerleri de kullanılarak veriye bir fit yapmamız gerekmekte.
#Burada bunun için Hesaplanan E değerleri kullanılacaktır!
#Sinusoidal bir data dağılımı olduğu OC Gateway sitesinden de anlaşılmaktadır.
#Burada hesaplanan E değerlerinin bulunduğu listeyi "array"e çevirmemiz gereklidir.
Evalues_calc=np.asarray(Evalues_calc)

#Burada parabolik fite kıyasla ekstradan bir terim gelmektedir.
#Bu terimde "Gerçek Ayrıklık" gerektiğinden E Dış Ayrıklık değerinden Nü Gerçek Ayrıklık değerlerini de elde etmemiz şarttır
#Şimdi bunun üzerinde çalışıyoruz.
Mean_Anomaly_values=[] #Ortalama Ayrıklık (M) değerleri
Eccentric_Anomaly_values=[] #Dış Ayrıklık (E) değerleri
True_Anomaly_values=[] #Gerçek Ayrıklık (ν) Değerleri

#M=2π/P*(t-T0) Ortalama Anomali

#E+esinE'deki E olan Dış Ayrıklık hesabı için iterasyon yapmamız gerekir.
#Bunun için E_ilk=M yaklaşımı yapılır.
#Buradan da E_yeni=M+e*sinE_eski formülü ile iterasyon gerçekleştirilir.

#E değerleri elde edildikten sonra tan(ν/2)=(√1+e/1-e)*tan(E/2) formülü kullanılarak Gerçek Anomali bulunur.
#Yukarıdaki formülden ν çekilerek yine E değerleri yerine yazılır ve Gerçek Anomali değerleri elde edilir.

#Unutulmamalıdır ki math kutüphanesindeki sin, cos, tan gibi trigonometrik parametreler "radyan" değerlerini kabul etmektedir.
#Bu nedenle trigonometrik işlemlerde içerinin radyan olması gerekmekte olup, bu unutulmamalıdır!
e=0.8 #Basıklık Değeri
omega=45 #Enberinin Argümanı Değeri

for i in range(len(Evalues_calc)):
    Mean_Anomaly=((2*math.pi)/P)*(OCvalues_calc[i])
    Mean_Anomaly_values.append(math.degrees(Mean_Anomaly))
       
    Ei_old=Mean_Anomaly
    Ei_new=Mean_Anomaly+(e*math.sin(Ei_old))
    error=abs(math.degrees(Ei_new)-math.degrees(Ei_old))
    while error>0.001:
        Ei_new=Mean_Anomaly+(e*math.sin(Ei_old))
        error=abs(math.degrees(Ei_new)-math.degrees(Ei_old))
        Ei_old=Ei_new
           
    Eccentric_Anomaly_values.append(math.degrees(Ei_new))
       
    True_Anomaly_0=90-omega
       
    True_Anomaly=2*math.atan(math.sqrt((1+e)/(1-e))*math.tan((math.radians(Eccentric_Anomaly_values[i])))/2)
    #True_Anomaly=math.degrees(True_Anomaly) #Derece olarak kaydetmek istersek bunu kullanıyoruz.
    if True_Anomaly<0:
        True_Anomaly=(2*math.pi)+True_Anomaly
    
    True_Anomaly_values.append(True_Anomaly)

#Ayrıca sinüs fonksiyonumuzun genliğini de aşağıdaki gibi tanımlayabiliriz.
A=abs((max(OCvalues_calc)-min(OCvalues_calc)))/2

#Burada hesaplanan ν değerlerinin bulunduğu listeyi de "array"e çevirmemiz gereklidir.
True_Anomaly_values=np.asarray(True_Anomaly_values)

#Şimdi Sinusoidal Fonksiyonu tanımlıyoruz.
def Sinusoidal(Evalues_calc,DeltaT0,DeltaP,Q,A,e,omega,True_Anomaly_values):
    part1=DeltaT0+(DeltaP*Evalues_calc)+(pow(Evalues_calc,2)*Q)
    part2=A/(1-(pow(e,2)*pow(math.cos(math.radians(omega)),2)))
    part3=3
    #part3=(((1-pow(e,2))/(1+(e*math.cos(math.radians(omega)))))*math.sin(math.radians((True_Anomaly_values+omega))))+(e*math.sin(math.radians(omega)))
    y=part1+(part2*part3)
    return y

#Şimdi de bu dataya eğri fitini yapıyoruz.
#Burada eğri fiti fonksiyonu 3 tane girdi istemektedir.
#Aynı şekilde eğri fiti de 2 tane çıktı vermektedir.
#Birinci çıktı parametreler için en uygun değerleri verir.
#İkinci çıktı ise parametrelerin hatalarını hesaplamak için kullanılan tahmini olasılıklardan oluşan matristir.
parameters_fit, covariance_fit = curve_fit(Sinusoidal, Evalues_calc, OCvalues_calc)

#Artık T0,P,Q parametrelerini yazdırabiliriz.
#Burada T0 aslında delta T0 olup T0'daki değişimi gösterir.
#Aynı şekilde P aslında delta P olup P'deki değişimi gösterir.
DeltaT0=parameters_fit[0]
DeltaP=parameters_fit[1]
Q=parameters_fit[2]
e=parameters_fit[3]
omega=parameters_fit[4]

#Bulduğumuz parametrelerin hatalarını da aşağıdaki gibi bulabiliriz.
#Hatayı hesaplamak için de matristeki köşegen değerlerin karekökü alınır.
error=np.sqrt(np.diag(covariance_fit))
DeltaT0_error=error[0]
DeltaP_error=error[1]
Q_error=error[2]
e_error=error[3]
omega_error=error[4]

#Yeni P ve T0 Değerlerimiz de aşağıdaki gibi olacaktır.
P=P+DeltaP
T0=T0+DeltaT0

#Şimdi de fitimizin uygunluğunu görmek için parametrelerini belirlediğimiz denklemi kullanıyoruz.
#Buradan y değerlerini hesaplıyoruz.
#Sonrasında da datamızı bu fit ile karşılaştırıyoruz.
#Ayrıca (O-C)'nin 0 olduğu noktadan bir çizgi çizersek eğrimizin şeklinden Basıklık değeri yaklaşık olarak bulunabilir.
zero_list=[]
for i in Evalues_calc:
    zero_list.append(0)

plt.plot(Evalues_calc,zero_list,'--', color="blue")
#Şimdi E-(O-C) grafiğini çizdirebiliriz.
plt.plot(Evalues_calc, OCvalues_calc, 'o', color="red")
#plt.plot(Evalues_calc, OCvalues_calc, 'o', label="Data")
fit_sinusoidal=Sinusoidal(Evalues_calc,DeltaT0,DeltaP,Q,A,e,omega,True_Anomaly_values)
#plt.plot(Evalues_calc,fit_sinusoidal, '-', label='Parabolic Fit',color='red')
#plt.legend(loc='upper right')
plt.xlabel('E (Çevrim) Değerleri')
plt.ylabel('O-C Değerleri')
plt.title('E-(O-C) Grafiği')
plt.show()

sys.exit()

#Burada önemli bir durum söz konusudur.
#Grafikte aşırı sapmaya uğrayan noktalar varsa bunların Excel Data Dosyasından silinmesi gerekir!
#Sonrasında tekrardan programın çalıştırılması uygun olacaktır, bu kontrol yapılmalıdır!

#Fitten Hesaplanan O-C değerlerini de aşağıdaki şekilde bir listeye atabiliriz.
OCvalues_fit=[]
for i in range(0,len(OCvalues_obs)):
    OC_fit_value=fit_sinusoidal[i]
    OCvalues_fit.append(OC_fit_value)
    
#Bu O-C değerlerini Excel Dosyasına kaydediyoruz.
#Fit üzerinden hesaplanan O-C değerleri artık bizim yeni "Hesaplanan O-C Değerlerimiz" olacaktır.
for i in range(len(OCvalues_fit)):
    datafile.cell(row=1,column=27).value="Fit ile Hesaplanan O-C Değerleri" #Başlık
    datafile.cell(row=i+2, column=27).value=str(round(OCvalues_fit[i],5)) #Değerler

#Şimdi Fitten Hesaplanan ve Webden Alınan O-C değerleri kullanılarak Ki Kare Hesabı yapılır.
chi_squared=0
for i in range(len(OCvalues_obs)):
    chi_squared+=pow((OCvalues_obs[i-1]-OCvalues_fit[i-1]),2)/(OCvalues_fit[i-1])

#Hesaplanan Ki Kare değerini de excel dosyasına atabiliriz.
datafile.cell(row=1,column=28).value="Ki Kare Değeri"
datafile.cell(row=2, column=28).value=str(chi_squared)

#Excel Dosyası kaydedilir.
datafile_open.save(r"D:\UBT 4. SINIF\UBT\Uzay Fiziğinde Seçilmiş Konular\Ödevler\V0410 Aur.xlsx")