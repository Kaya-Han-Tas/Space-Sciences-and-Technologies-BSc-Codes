from scipy.optimize import curve_fit
from matplotlib import pyplot as plt
import math
import numpy as np
import openpyxl

#Excel Dosyasından data çekilecektir
#Açtığımız andaki dosya "Workbook" olarak geçer ve üzerine yazılabilir.
#Excel dosyamıza yazabilmek için Dosyanın kapatılması gerekmekte olup, açık olmamalıdır! 
#Aksi takdirde hata alınacaktır.
datafile_open=openpyxl.load_workbook(r"D:\UBT 4. SINIF\UBT\Uzay Fiziğinde Seçilmiş Konular\Ödevler\CW Peg Fix 2.xlsx")

#.active ile açılan dosya bir "active sheet" haline getirilir ve buradan datalar çekilebilir.
datafile=datafile_open.active

#Data Dosyamızdan kullanılacak T0 ve P değeri çekilir.
T0=float((datafile.cell(row=2,column=4).value)) #Başlangıç Zamanı
P=float((datafile.cell(row=2,column=5).value)) #Periyot

#Buradan itibaren webden alınan değerler "obs" ve hesaplanan değerler "calc" ile gösterilmektedir.
#Bununla beraber diğer Dataları da çekerek kendi listelerine aktarıyoruz. (Webden Alınan Veriler)
Evalues_obs=[]
Tvalues_obs=[]
OCvalues_obs=[]
x=(datafile.max_row)-1
y=1

while y<=x:
    y+=1
    Edata=(datafile.cell(row=y,column=1).value)
    Evalues_obs.append(Edata)
    OCdata=(datafile.cell(row=y,column=2).value)
    OCvalues_obs.append(OCdata)
    Tdata=(datafile.cell(row=y,column=3).value)
    Tvalues_obs.append(Tdata)

#T değerleri ve OC değerleri Excel Dosyasında "Metin" olarak yazıldığından bunları sayıya çevirmemiz gerekir.
Tvalues_obs=[float (i) for i in Tvalues_obs]
OCvalues_obs=[float(i) for i in OCvalues_obs]

#Şimdi datalarımızı array'e çevirerek işimizi kolaylaştırabiliriz.
Tvalues_obs=np.asarray(Tvalues_obs)
Evalues_obs=np.asarray(Evalues_obs)
OCvalues_obs=np.asarray(OCvalues_obs)

#İlk olarak kendimiz E değerlerini hesaplıyoruz.
#Bunun için E=(T-T0)/P formülü kullanılacaktır.
#Buna dikkat edilerek işlem gerçekleştirilir.
#Ayrıca E değerlerinin yuvarlanmasında bir kural üzerinden ilerlememiz gerekir.
#Bunun için bulduğumuz Hesaplanan "Yuvarlanmış" E değeri ile Hesaplanan E değeri arasındaki farkın mutlak değerinin 0.3'ten büyük olduğunu varsayıyoruz.

#Hesaplanan "Yuvarlanmış" Değer, Hesaplanan Değerden küçükse yani ikisinin farkı negatifse Yuvarlanan değerden 0.5 ÇIKARILIR.
#Hesaplanan "Yuvarlanmış" Değer, Hesaplanan Değerden büyükse yani ikisinin farkı pozitifse Yuvarlanan değere 0.5 EKLENİR.

#Hesaplanan "Yuvarlanmış" E değeri ile Hesaplanan E değeri arasındaki farkın mutlak değerinin 0.3'ten küçük olması durumunda yuvarlanan değer doğru kabul edilir.
#Buradan yola çıkarak E değerlerini doğru yuvarlanmış biçimleriyle elde ederiz.
Evalues_calc=[]
Evalues_calc_original=[]
for i in range(len(Evalues_obs)):
    E_calc=(Tvalues_obs[i]-T0)/P
    Evalues_calc_original.append(E_calc)
    
    if abs(E_calc-round(E_calc))>0.3:
        if (E_calc-round(E_calc))<0:
            Evalues_calc.append(round(E_calc)-0.5)
        
        else:
            Evalues_calc.append(round(E_calc)+0.5)
    
    else:
        Evalues_calc.append(round(E_calc))

#Şimdi bulunan E değerlerini Excel Dosyamıza yazmamız faydalı olacaktır.
#Bunun için de aşağıdaki kod kullanılır.
#Burada artı olarak bulunan değerleri "Metin" olarak Excel dosyasına yazmak faydalı olacaktır.

#i=0 olacağından ve 1. satırda başlık bulunduğundan sayılar excel dosyasına 2. satırdan itibaren yazılmalıdır.
#Bu nedenle i+2. satırdan itibaren değerleri Excel dosyasına yazmaktayız.
#Ayrıca E değerlerinde E değeri x.50 veya üzeri gibi bir değer çıkması durumunda bu secondary minimuma ait olup 0.5 eklenerek E değeri alınmalıdır.
for i in range(len(Evalues_calc)):
    datafile.cell(row=1,column=8).value="Hesaplanan E Değerleri" #Başlık
    datafile.cell(row=i+2, column=8).value=(Evalues_calc[i])

#Siteden Alınan E değerleri ile Hesaplanan E değerleri arası farkı da Excel Dosyasına yazdırarak kontrol yapabiliriz.
#Burada "Secondary" tutulmalardan elde edilen E değerleri arası fark çıkacaktır, buradan Secondary tutulma olduğu anlaşılabilir!
for i in range(len(Evalues_calc)):
    datafile.cell(row=1,column=9).value="Site ve Hesaplanan E değeri arası fark" #Başlık
    datafile.cell(row=i+2, column=9).value=(Evalues_obs[i]-Evalues_calc[i]) #Değerler

#Hesaplanan E Değerlerinin yuvarlanmamış durumlarını da kontrol için Excel Dosyasına yazdırıyoruz.
for i in range(len(Evalues_calc)):
    datafile.cell(row=1,column=10).value="Hesaplanan E Değerleri (Yuvarlanmamış)" #Başlık
    datafile.cell(row=i+2, column=10).value=(Evalues_calc_original[i]) #Değerler

#İkinci olarak O-C Hesabımızı yapıyoruz.
#Bunun için T=T0+EP formülü kullanılacaktır.
#Bu formülde bulunan E değerleri Hesapladığımız E değerleridir!

#Bu formülden elde edilen değerler "Calculated" Zamanlardır.
#Siteden aldıklarımız ise "Observed" Zamanlardır.
#Bu iki zaman farkı da bize O-C değerlerini vermelidir.

#İlk olarak Calculated yani Hesaplanan Zaman değerlerini T=T0+EP formülü ile elde ediyoruz.
#Burada T0 değeri primary veya secondary tutulma olmasına göre belirlenmelidir!
#Bunu dikkate alarak Bulduğumuz T değerlerini ayrıca Excel Dosyasına yazdırıyoruz.
Tvalues_calc=[]
for i in range(len(Tvalues_obs)):
    T_calc=T0+(Evalues_calc[i]*P)
    Tvalues_calc.append(T_calc)
        
    datafile.cell(row=1,column=13).value="Hesaplanan T Değerleri (Calculated Time)" #Başlık
    datafile.cell(row=i+2, column=13).value=(T_calc) #Değerler

#Hesapladığımız Zaman değerleri ve Websitesinden aldığımız zaman değerlerini kullanarak O-C hesaplanır.
#Excel Dosyasına bu değer de yazdırılır.
OCvalues_calc=[]
for i in range(len(Tvalues_obs)):
    OC_calc=Tvalues_obs[i]-Tvalues_calc[i]
    OCvalues_calc.append(OC_calc)
    datafile.cell(row=1,column=14).value="Hesaplanan O-C Değerleri" #Başlık
    datafile.cell(row=i+2, column=14).value=(OC_calc) #Değerler
    
#Şimdi bu O-C değerlerinin Sitedeki O-C değerleri ile karşılaştırmasını Excel Dosyamıza yazmamız faydalı olacaktır.
#Bunun için de aşağıdaki kod kullanılır.
#Burada artı olarak bulunan değerleri "Metin" olarak Excel dosyasına yazmak faydalı olacaktır.

#i=0 olacağından ve 1. satırda başlık bulunduğundan sayılar excel dosyasına 2. satırdan itibaren yazılmalıdır.
#Bu nedenle i+2. satırdan itibaren değerleri Excel dosyasına yazmaktayız.
for i in range(len(OCvalues_calc)):
    datafile.cell(row=1,column=17).value="Hesaplanan O-C Değerleri" #Başlık
    datafile.cell(row=i+2, column=17).value=str(round(OCvalues_calc[i],5)) #Değerler

#Yine E için yaptığımız gibi Siteden alınan O-C değerleri ile Hesapladığımız O-C değerleri arası farkı Excel Dosyasına yazdırıyoruz.
for i in range(len(Evalues_calc)):
    datafile.cell(row=1,column=18).value="Site ve Hesaplanan O-C değeri arası fark" #Başlık
    datafile.cell(row=i+2, column=18).value=(round(OCvalues_calc[i],5)-OCvalues_obs[i]) #Değerler

#Şimdi bulduğumuz O-C değerleri de kullanılarak veriye bir fit yapmamız gerekmekte.
#Burada bunun için Hesaplanan E değerleri kullanılacaktır!
#Parabolik bir data dağılımı olduğu OC Gateway sitesinden de anlaşılmaktadır.
#Burada hesaplanan E değerlerinin bulunduğu listeyi "array"e çevirmemiz gereklidir.
Evalues_calc=np.asarray(Evalues_calc)

#Önce Parabolik Fonksiyonu tanımlıyoruz.
def Parabolic(Evalues_calc,T0,P,Q):
    y=T0+(P*Evalues_calc)+(pow(Evalues_calc,2)*Q)
    return y

#Şimdi de bu dataya eğri fitini yapıyoruz.
#Burada eğri fiti fonksiyonu 3 tane girdi istemektedir.
#Aynı şekilde eğri fiti de 2 tane çıktı vermektedir.
#Birinci çıktı parametreler için en uygun değerleri verir.
#İkinci çıktı ise parametrelerin hatalarını hesaplamak için kullanılan tahmini olasılıklardan oluşan matristir.
parameters_fit, covariance_fit = curve_fit(Parabolic, Evalues_calc, OCvalues_calc)

#Artık T0,P,Q parametrelerini yazdırabiliriz.
#Burada T0 aslında delta T0 olup T0'daki değişimi gösterir.
#Aynı şekilde P aslında delta P olup P'deki değişimi gösterir.
T0_value_fit=parameters_fit[0]
P_value_fit=parameters_fit[1]
Q_value_fit=parameters_fit[2]

#Bunları da Excel Dosyasına yazabiliriz.
datafile.cell(row=1,column=20).value="Fitten Elde Edilen Delta T0 değeri"
datafile.cell(row=2, column=20).value=str(T0_value_fit)

datafile.cell(row=1,column=21).value="Fitten Elde Edilen Delta P değeri"
datafile.cell(row=2, column=21).value=str(P_value_fit)

datafile.cell(row=1,column=22).value="Fitten Elde Edilen Q değeri"
datafile.cell(row=2, column=22).value=str(Q_value_fit)

#Bulduğumuz parametrelerin hatalarını da aşağıdaki gibi bulabiliriz.
#Hatayı hesaplamak için de matristeki köşegen değerlerin karekökü alınır.
error=np.sqrt(np.diag(covariance_fit))
T0_value_fit_error=error[0]
P_value_fit_error=error[1]
Q_value_fit_error=error[2]

#Bunları da Excel Dosyasına yazıyoruz.
datafile.cell(row=4,column=20).value="Fitten Elde Edilen Delta T0 değerinin Hatası"
datafile.cell(row=5, column=20).value=str(T0_value_fit_error)

datafile.cell(row=4,column=21).value="Fitten Elde Edilen Delta P değerinin Hatası"
datafile.cell(row=5, column=21).value=str(P_value_fit_error)

datafile.cell(row=4,column=22).value="Fitten Elde Edilen Q değerinin Hatası"
datafile.cell(row=5, column=22).value=str(Q_value_fit_error)

#Şimdi de fitimizin uygunluğunu görmek için parametrelerini belirlediğimiz denklemi kullanıyoruz.
#Buradan y değerlerini hesaplıyoruz.
#Sonrasında da datamızı bu fit ile karşılaştırıyoruz.
#Şimdi E-(O-C) grafiğini çizdirebiliriz.
plt.plot(Evalues_calc, OCvalues_calc, 'o', label="Data")
fit_parabolic=Parabolic(Evalues_calc,T0_value_fit,P_value_fit,Q_value_fit)
plt.plot(Evalues_calc,fit_parabolic, '-', label='Parabolic Fit',color='red')
plt.legend(loc='upper right')
plt.xlabel('E (Çevrim) Değerleri')
plt.ylabel('O-C Değerleri')
plt.title('E-(O-C) Grafiği')
plt.show()

#Burada önemli bir durum söz konusudur.
#Grafikte aşırı sapmaya uğrayan noktalar varsa bunların Excel Data Dosyasından silinmesi gerekir!
#Sonrasında tekrardan programın çalıştırılması uygun olacaktır, bu kontrol yapılmalıdır!

#Fitten Hesaplanan O-C değerlerini de aşağıdaki şekilde bir listeye atabiliriz.
OCvalues_fit=[]
for i in range(0,len(OCvalues_obs)):
    OC_fit_value=fit_parabolic[i]
    OCvalues_fit.append(OC_fit_value)
    
#Bu O-C değerlerini Excel Dosyasına kaydediyoruz.
#Fit üzerinden hesaplanan O-C değerleri artık bizim yeni "Hesaplanan O-C Değerlerimiz" olacaktır.
for i in range(len(OCvalues_fit)):
    datafile.cell(row=1,column=27).value="Fit ile Hesaplanan O-C Değerleri" #Başlık
    datafile.cell(row=i+2, column=27).value=str(round(OCvalues_fit[i],5)) #Değerler

#Şimdi Fitten Hesaplanan ve Webden Alınan O-C değerleri kullanılarak Ki Kare Hesabı yapılır.
chi_squared=0
for i in range(len(OCvalues_obs)):
    chi_squared+=pow((OCvalues_obs[i-1]-OCvalues_fit[i-1]),2)/(OCvalues_fit[i-1])

#Hesaplanan Ki Kare değerini de excel dosyasına atabiliriz.
datafile.cell(row=1,column=28).value="Ki Kare Değeri"
datafile.cell(row=2, column=28).value=str(chi_squared)

#Excel Dosyası kaydedilir.
datafile_open.save(r"D:\UBT 4. SINIF\UBT\Uzay Fiziğinde Seçilmiş Konular\Ödevler\CW Peg Test 2.xlsx")

#Buraya kadar yapılanları tekrar yapıyoruz.
#0. Adım: E değerleri yeni T0 ve P değerlerine göre tekrardan hesaplanır.
#1. Adım: T=T0+EP formülünden Calculated Zamanı hesaplamak için fitten elde ettiğimiz yeni T0 ve P değerleri kullanılır ve T Calculatedler elde edilir.
#2. Adım: Webden elde edilen T değerleri ile Hesaplanan T değerleri farkından O-C değerleri elde edilir. (O-C Calculated)
#3. Adım: O-C ve E değerleri kullanılarak yeni bir fit yapılır. Buradan yeni T0, P ve Q değerleri elde edilir.
#4. Adım: Elde edilen T0, P ve Q değerleri O-C=E^2Q+EP+T0 formülünde yerine yazılır. Buradan yeni O-C değerleri elde edilir. (O-C Fit)
#5. Adım: Ki Kare hesabı yapılır.
#6. Adım: Ki Kare değeri azalıyorsa 1. Adımdan bu işlem tekrar yapılır. Ki Kare değeri azalmıyorsa işlem bitirilir.
#7. Adım: En son olarak elde edilen T0, P, Q ve O-C değerleri Excel Dosyasına kaydedilir.

#Bu adımlardan yola çıkarak işlemlerimizi döngü içerisinde gerçekleştiriyoruz.
chi_squared_new=chi_squared
chi_squared_old=chi_squared-0.001

T0=T0+T0_value_fit
P=P+P_value_fit
Q=Q_value_fit

while chi_squared_new<chi_squared_old:
    #0. Adım
    Evalues_calc=[]
    for i in range(len(Evalues_obs)):
        E_calc=(Tvalues_obs[i]-T0)/P
        Evalues_calc_original.append(E_calc)
    
        if abs(E_calc-round(E_calc))>0.3:
            if (E_calc-round(E_calc))<0:
                Evalues_calc.append(round(E_calc)-0.5)
        
            else:
                Evalues_calc.append(round(E_calc)+0.5)
    
        else:
            Evalues_calc.append(round(E_calc))
    
    Evalues_calc=np.asarray(Evalues_calc)
    
    #1. Adım
    Tvalues_calc=[]
    for i in range(len(Tvalues_obs)):
        T_calc=T0+(Evalues_calc[i]*P)
        Tvalues_calc.append(T_calc)
    
    #2. Adım
    OCvalues_calc=[]
    for i in range(len(Tvalues_obs)):
        OC_calc=Tvalues_obs[i]-Tvalues_calc[i]
        OCvalues_calc.append(OC_calc)

    #3. Adım
    def Parabolic_Loop(Evalues_calc,T0,P,Q):
        y=T0+(P*Evalues_calc)+(pow(Evalues_calc,2)*Q)
        return y
    
    parameters_fit, covariance_fit = curve_fit(Parabolic_Loop, Evalues_calc, OCvalues_calc)
    
    T0_value_fit=parameters_fit[0]
    P_value_fit=parameters_fit[1]
    Q_value_fit=parameters_fit[2]
    
    error=np.sqrt(np.diag(covariance_fit))
    T0_value_fit_error=error[0]
    P_value_fit_error=error[1]
    Q_value_fit_error=error[2]
    
    #4. Adım
    T0=T0+T0_value_fit
    P=P+P_value_fit
    Q=Q_value_fit
    
    fit_parabolic=Parabolic(Evalues_calc,T0_value_fit,P_value_fit,Q_value_fit)
    OCvalues_fit=[]
    for i in range(0,len(OCvalues_obs)):
        OC_fit_value=fit_parabolic[i]
        OCvalues_fit.append(OC_fit_value)
    
    print(OCvalues_calc)
    print(OCvalues_fit)
    
    #5. Adım
    chi_squared_calc=0
    for i in range(len(OCvalues_obs)):
        chi_squared_calc+=pow((OCvalues_obs[i-1]-OCvalues_fit[i-1]),2)/(OCvalues_fit[i-1])
    
    #6. Adım
    chi_squared_old=chi_squared_new
    chi_squared_new=chi_squared_calc
    
    print(chi_squared_old)
    print(chi_squared_new)
    print(50*"-")
    
#Artık istenilen değerler elde edildiğine göre direkt olarak Excel'e bu değerler kayıt edilir.
for i in range(len(OCvalues_fit)):
    datafile.cell(row=1,column=27).value="Fit ile Hesaplanan O-C Değerleri" #Başlık
    datafile.cell(row=i+2, column=27).value=str(round(OCvalues_fit[i],5)) #Değerler

datafile.cell(row=1,column=20).value="Fitten Elde Edilen Delta T0 değeri"
datafile.cell(row=2, column=20).value=str(T0_value_fit)

datafile.cell(row=1,column=21).value="Fitten Elde Edilen Delta P değeri"
datafile.cell(row=2, column=21).value=str(P_value_fit)

datafile.cell(row=1,column=22).value="Fitten Elde Edilen Q değeri"
datafile.cell(row=2, column=22).value=str(Q_value_fit)

datafile.cell(row=4,column=20).value="Fitten Elde Edilen Delta T0 değerinin Hatası"
datafile.cell(row=5, column=20).value=str(T0_value_fit_error)

datafile.cell(row=4,column=21).value="Fitten Elde Edilen Delta P değerinin Hatası"
datafile.cell(row=5, column=21).value=str(P_value_fit_error)

datafile.cell(row=4,column=22).value="Fitten Elde Edilen Q değerinin Hatası"
datafile.cell(row=5, column=22).value=str(Q_value_fit_error)

datafile.cell(row=7,column=20).value="Fitten Elde Edilen T0 değeri"
datafile.cell(row=8, column=20).value=str(T0)

datafile.cell(row=7,column=21).value="Fitten Elde Edilen P değeri"
datafile.cell(row=8, column=21).value=str(P)

datafile.cell(row=1,column=28).value="Ki Kare Değeri"
datafile.cell(row=2, column=28).value=str(chi_squared_new)

#Excel Dosyası kaydedilir.
datafile_open.save(r"D:\UBT 4. SINIF\UBT\Uzay Fiziğinde Seçilmiş Konular\Ödevler\CW Peg Fix 2.xlsx")

#Artık düzeltilmiş T0,P, Q ve O-C değerleri ile grafiğimizi yeniden çizdirebiliriz.
plt.plot(Evalues_calc, OCvalues_calc, 'o', label="Data")
fit_parabolic=Parabolic(Evalues_calc,T0_value_fit,P_value_fit,Q_value_fit)
plt.plot(Evalues_calc,fit_parabolic, '-', label='Parabolic Fit',color='red')
plt.legend(loc='upper right')
plt.xlabel('E (Çevrim) Değerleri')
plt.ylabel('O-C Değerleri')
plt.title('Düzeltilmiş E-(O-C) Grafiği')
plt.show()