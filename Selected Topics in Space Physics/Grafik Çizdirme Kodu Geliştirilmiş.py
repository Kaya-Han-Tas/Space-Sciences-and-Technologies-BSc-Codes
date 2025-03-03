from scipy.optimize import curve_fit
from matplotlib import pyplot as plt
import math
import numpy as np
import openpyxl

#Excel Dosyasından data çekilecektir
#Açtığımız andaki dosya "Workbook" olarak geçer ve üzerine yazılabilir.
#Excel dosyamıza yazabilmek için Dosyanın kapatılması gerekmekte olup, açık olmamalıdır! 
#Aksi takdirde hata alınacaktır.
datafile_open=openpyxl.load_workbook(r"D:\UBT 4. SINIF\UBT\Uzay Fiziğinde Seçilmiş Konular\Ödevler\CW Peg.xlsx")

#.active ile açılan dosya bir "active sheet" haline getirilir ve buradan datalar çekilebilir.
datafile=datafile_open.active

#Data Dosyamızdan kullanılacak T0 ve P değeri çekilir.
T0=float((datafile.cell(row=2,column=4).value)) #Başlangıç Zamanı
P=float((datafile.cell(row=2,column=5).value)) #Periyot

#Diğer Dataları da çekerek kendi listelerine aktarıyoruz.
Evalues=[]
Tvalues_obs=[]
OCvalues=[]
x=(datafile.max_row)-1
y=1
Edata=0
Tdata=0

while y<=x:
    y+=1
    Edata=(datafile.cell(row=y,column=1).value)
    Evalues.append(Edata)
    OCdata=(datafile.cell(row=y,column=2).value)
    OCvalues.append(OCdata)
    Tdata=(datafile.cell(row=y,column=3).value)
    Tvalues_obs.append(Tdata)

#T değerleri ve OC değerleri Excel Dosyasında "Metin" olarak yazıldığından bunları sayıya çevirmemiz gerekir.
Tvalues_obs=[float (i) for i in Tvalues_obs]
OCvalues=[float(i) for i in OCvalues]

#Şimdi datalarımızı array'e çevirerek işimizi kolaylaştırabiliriz.
Tvalues_obs=np.asarray(Tvalues_obs)
Evalues=np.asarray(Evalues)
OCvalues=np.asarray(OCvalues)

#İlk olarak O-C Hesabımızı yapıyoruz.
#Bunun için T=T0+EP formülü kullanılacaktır.
#Bu formülden elde edilen değerler "Calculated" Zamanlardır.
#Siteden aldıklarımız ise "Observed" Zamanlardır.
#Bu iki zaman farkı da bize O-C değerlerini vermelidir.

#İlk olarak Calculated yani Hesaplanan Zaman değerlerini T=T0+EP formülü ile elde ediyoruz.
Tvalues_calc=[]
for i in range(len(Tvalues_obs)):
    T_calc=T0+(Evalues[i]*P)
    Tvalues_calc.append(T_calc)

#Hesapladığımız Zaman değerleri ve Websitesinden aldığımız zaman değerlerini kullanarak O-C hesaplanır.
OCvalues_calc=[]
for i in range(len(Tvalues_obs)):
    OC_calc=Tvalues_obs[i]-Tvalues_calc[i]
    OCvalues_calc.append(OC_calc)
    
#Bu O-C değerlerini Excel Dosyamıza yazmamız faydalı olacaktır.
#Bunun için de aşağıdaki kod kullanılır.
#Burada artı olarak bulunan değerleri "Metin" olarak Excel dosyasına yazmak faydalı olacaktır.
z=1 #İterasyon Sayısı
c=8 #Sütun Sayısı

#i=0 olacağından ve 1. satırda başlık bulunduğundan sayılar excel dosyasına 2. satırdan itibaren yazılmalıdır.
#Bu nedenle i+2. satırdan itibaren değerleri Excel dosyasına yazmaktayız.
for i in range(len(OCvalues_calc)):
    datafile.cell(row=1,column=c).value="Hesaplanan O-C Değerleri" #Başlık
    datafile.cell(row=i+2, column=c).value=str(round(OCvalues_calc[i],5)) #Değerler

#Şimdi bulduğumuz O-C değerleri de kullanılarak veriye bir fit yapmamız gerekmekte.
#Parabolik bir data dağılımı olduğu anlaşılmaktadır.
#Önce Parabolik Fonksiyonu tanımlıyoruz.
def Parabolic(Evalues,T0,P,Q):
    y=T0+(P*Evalues)+(pow(Evalues,2)*Q)
    return y

#Şimdi de bu dataya eğri fitini yapıyoruz.
#Burada eğri fiti fonksiyonu 3 tane girdi istemektedir.
#Aynı şekilde eğri fiti de 2 tane çıktı vermektedir.
#Birinci çıktı parametreler için en uygun değerleri verir.
#İkinci çıktı ise parametrelerin hatalarını hesaplamak için kullanılan tahmini olasılıklardan oluşan matristir.
parameters_fit, covariance_fit = curve_fit(Parabolic, Evalues, OCvalues_calc)

#Artık T0,P,Q parametrelerini yazdırabiliriz.
#Burada T0 aslında delta T0 olup T0'daki değişimi gösterir.
#Aynı şekilde P aslında delta P olup P'deki değişimi gösterir.
T0_value_fit=parameters_fit[0]
P_value_fit=parameters_fit[1]
Q_value_fit=parameters_fit[2]

print("Parabolik Fit: ", 50*"-")
print("Delta T0 parametresinin değeri: ", T0_value_fit)
print("Delta P parametresinin değeri: ", P_value_fit)
print("Q parametresinin değeri: ", Q_value_fit)
print(60*"-")

#Bunları da Excel Dosyasına yazabiliriz.
datafile.cell(row=1,column=c+1).value="Fitten Elde Edilen T0 değeri"
datafile.cell(row=2, column=c+1).value=str(T0_value_fit)

datafile.cell(row=1,column=c+2).value="Fitten Elde Edilen P değeri"
datafile.cell(row=2, column=c+2).value=str(P_value_fit)

datafile.cell(row=1,column=c+3).value="Fitten Elde Edilen Q değeri"
datafile.cell(row=2, column=c+3).value=str(Q_value_fit)

#Bulduğumuz parametrelerin hatalarını da aşağıdaki gibi bulabiliriz.
#Hatayı hesaplamak için de matristeki köşegen değerlerin karekökü alınır.
error=np.sqrt(np.diag(covariance_fit))
T0_value_fit_error=error[0]
P_value_fit_error=error[1]
Q_value_fit_error=error[2]

print("Delta T0 parametresinin hatası: ",T0_value_fit_error)
print("Delta P parametresinin hatası: ",P_value_fit_error)
print("Q parametresinin hatası: ",Q_value_fit_error)
print(60*"-")

#Bunları da Excel Dosyasına yazıyoruz.
datafile.cell(row=4,column=c+1).value="Fitten Elde Edilen T0 değerinin Hatası"
datafile.cell(row=5, column=c+1).value=str(T0_value_fit_error)

datafile.cell(row=4,column=c+2).value="Fitten Elde Edilen P değerinin Hatası"
datafile.cell(row=5, column=c+2).value=str(P_value_fit_error)

datafile.cell(row=4,column=c+3).value="Fitten Elde Edilen Q değerinin Hatası"
datafile.cell(row=5, column=c+3).value=str(Q_value_fit_error)

#Excel Dosyası kaydedilir.
datafile_open.save(r"D:\UBT 4. SINIF\UBT\Uzay Fiziğinde Seçilmiş Konular\Ödevler\CW Peg.xlsx")

#Şimdi de fitimizin uygunluğunu görmek için parametrelerini belirlediğimiz denklemi kullanıyoruz.
#Buradan y değerlerini hesaplıyoruz.
#Sonrasında da datamızı bu fit ile karşılaştırıyoruz.
#Şimdi E-(O-C) grafiğini çizdirebiliriz.
plt.plot(Evalues, OCvalues_calc, 'o', label="Data")
fit_parabolic=Parabolic(Evalues,T0_value_fit,P_value_fit,Q_value_fit)
plt.plot(Evalues,fit_parabolic, '-', label='Parabolic Fit',color='red')
plt.legend(loc='upper right')
plt.xlabel('E (Çevrim) Değerleri')
plt.ylabel('O-C Değerleri')
plt.title('E-(O-C) Grafiği')
plt.show()

#Hesaplanan O-C değerlerini de aşağıdaki şekilde bir listeye atabiliriz.
OC_fit_values=[]
for i in range(0,len(OCvalues)):
    OC_fit_value=fit_parabolic[i]
    OC_fit_values.append(OC_fit_value)

#Bağıl Hata ile Teorik (Hesaplanan) OC değerleri ile Deneysel (Siteden alınan) OC değerleri arası hatayı buluruz.
OC_Error=[]
for i in range(1,len(OCvalues)):
    OC_error_value=abs(OC_fit_values[i]-OCvalues[i])/(OC_fit_values[i])
    OC_Error.append(OC_error_value)

print("Hesaplanan ve Gözlemsel OC değerleri arası bağıl hata değeri: ")
OC_Error=np.asarray(OC_Error)
for i in range(0,len(OCvalues)-1):
    print(i+1,". OC hatası:",OC_Error[i])

print(60*"-")

#Şimdi de P-E grafiğini çizdiriyoruz.
#Burada Periyot Pyeni=Peski+2EQ'dan bulunur.
Pvalues=[]
Pold=2.372527
for i in range(0,len(Evalues)):
    Pnew=Pold+(2*Evalues[i]*Q_value_fit)
    Pvalues.append(Pnew)

#print(Pvalues)
#Yukarıdaki "#" silinerek bulunan P değerlerini elde edebiliriz.
#Artık E-P grafiğini çizebiliriz.
plt.plot(Evalues, Pvalues, 'o')
plt.xlabel('E (Çevrim) Değerleri')
plt.ylabel('P (Periyot) Değerleri')
plt.title('E-P Grafiği')
plt.show()